import logging
import openpyxl
import azure.functions as func

DEMANDE_COLUMN_NAMES= ['JOUR','SHIFT','DEPOT','CM','TYPE','DEMANDE']
FLOTTE_COLUMN_NAMES= ["IMMATRICULATION","TYPE","EXPLOITATION","CAPACITE","POIDS"]


def validate_excel_file(file, filename):
    # Check the file extension
    if not filename.endswith('.xlsx'):
        return False
    # Load the Excel file
    try:

        workbook = openpyxl.load_workbook(file)

        if "DEMANDE" in  filename:
            sheet_name= workbook.sheetnames[0]
            sheet = workbook[sheet_name]

            min_row = sheet.min_row
            max_row = sheet.max_row

            if min_row == max_row:
                logging.info("The file does not have any data rows.")
                return False

            column_names = []
            for cell in sheet[1]:
                if cell.value is None:
                    break
                column_names.append(cell.value)
            
            if column_names == DEMANDE_COLUMN_NAMES:
                logging.info("Column names are correct")
                return True


        if "FLOTTE" in  filename:
            sheet_name= workbook.sheetnames[0]
            sheet = workbook[sheet_name]

            dims = sheet.calculate_dimension()
            min_row = sheet.min_row
            max_row = sheet.max_row

            if min_row == max_row:
                logging.info("The file does not have any data rows.")
                return False

            column_names = []
            for cell in sheet[1]:
                if cell.value is None:
                    break
                column_names.append(cell.value)
            
            if column_names == FLOTTE_COLUMN_NAMES:
                print("Column names are correct")
                return True


    except openpyxl.utils.exceptions.InvalidFileException:
        return False


def main(req: func.HttpRequest) -> func.HttpResponse:

    logging.info('Python HTTP trigger function processed a request.')


    if 'fileName' not in req.files:
        return 'No file found in the request.'

    file = req.files['fileName']
    filename = file.filename

    if(validate_excel_file(file,filename)):
        return func.HttpResponse(f"The file {filename} is valid and can be processed")
    
    else:
        return func.HttpResponse(f"The file {filename} is not valid and can not be processed")
        
